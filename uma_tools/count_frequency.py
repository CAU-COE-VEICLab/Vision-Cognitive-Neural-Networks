# --------------------------------------------------------
# The potential of cognitive-inspired neural network modeling framework for computer vision processing tasks
# Licensed under The MIT License [see LICENSE for details]
# Written by Guorun Li
# --------------------------------------------------------

# You can calculate the SSIM value of each image in your dataset by following this step: 
# 1. Using 'statistic_uma_strategy1.py' & 'statistic_uma_strategy2.py' to calculate the SSIM value of each image in the dataset.
#    Then you can get the Excel file (named ssim_origin_excel_file), which contains the SSIM value of each image in your dataset 
# 2. Using 'count_frequency.py' to calculate the frequency distribution P in your dataset.

import openpyxl
import pandas as pd


def updataexcel(excel_path, data_dic, column_number):
    new_data = pd.DataFrame([data_dic])
    try:
        df = pd.read_excel(excel_path, engine='openpyxl')
        result = pd.concat([df, new_data], ignore_index=True)
        result.to_excel(excel_path, index=False, engine='openpyxl')

    except FileNotFoundError:
        dic_save = {}
        for column in range(column_number + 1):
            dic_save[column] = dic_save.get(column, [])
        df = pd.DataFrame(dic_save)
        print("Creat a new Excel file:", excel_path)
        result = pd.concat([df, new_data], ignore_index=True)
        result.to_excel(excel_path, index=False, engine='openpyxl')


def statistical_frequency(file_path, frequency_dic):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    index = 0
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        row_length = 0
        for cell_value in row:
            if cell_value is not None:
                row_length += 1
            else:
                row_length -= 1
                break
        index += 1

        frequency_dic[row_length] = frequency_dic.get(row_length, 0) + 1

        print(f"File name{file_path} Row {sheet.title}: Length of non-empty cells = {row_length}, index {index}")

    workbook.close()
    return frequency_dic, index

if __name__ == '__main__':
    # same with statistic_uma_strategy1.py & statistic_uma_strategy2.py
    target_size = 224
    file_path_strategy1 = 'the path where the ssim values counted by statistic_uma_strategy1.py (ssim_origin_excel_file.xlsx)'
    statical_strategy1_ssim_origin_excel_file = 'the path where the frequency distribution P are saved (file endwith .xlsx)'
    file_path_stategy2 = 'the path where the ssim values counted by statistic_uma_strategy2.py (ssim_origin_excel_file.xlsx)'
    statical_strategy2_ssim_origin_excel_file = 'the path where the frequency distribution P are saved (file endwith .xlsx)'

    fft2sampling_frequncy_dic = {}
    sampling2fft_frequncy_dic = {}
    index = 0
    fft2sampling_frequncy_dic, current_index = statistical_frequency(file_path_strategy1, fft2sampling_frequncy_dic)
    index += current_index

    updataexcel(excel_path=statical_strategy1_ssim_origin_excel_file, data_dic=fft2sampling_frequncy_dic, column_number=target_size // 2 + 1)
    print(f'file {statical_strategy1_ssim_origin_excel_file} , Number of searches {index}')

    index = 0
    sampling2fft_frequncy_dic, current_index = statistical_frequency(file_path_stategy2, sampling2fft_frequncy_dic)
    index += current_index
 
    updataexcel(excel_path=statical_strategy2_ssim_origin_excel_file, data_dic=sampling2fft_frequncy_dic, column_number=target_size // 4 + 1)
    print(f'file {statical_strategy2_ssim_origin_excel_file} , Number of searches {index}')